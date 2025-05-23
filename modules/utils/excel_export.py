"""
Módulo para exportación de datos a Excel con formato
"""

import io
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

def to_excel_with_format(tabla: pd.DataFrame) -> io.BytesIO:
    """
    Convierte una tabla a Excel con formato para resaltar fila y columna Total
    
    Args:
        tabla: DataFrame a convertir
        
    Returns:
        BytesIO con el archivo Excel formateado
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        tabla.to_excel(writer, sheet_name='Pendientes')
        ws = writer.sheets['Pendientes']
        
        # Resaltar fila Total
        total_row = ws.max_row
        for col in range(1, ws.max_column + 1):
            ws.cell(row=total_row, column=col).font = Font(bold=True)
            ws.cell(row=total_row, column=col).fill = PatternFill(
                start_color='FFFF00', end_color='FFFF00', fill_type='solid'
            )
        
        # Resaltar columna Total
        total_col = ws.max_column
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=total_col).font = Font(bold=True)
            ws.cell(row=row, column=total_col).fill = PatternFill(
                start_color='FFFF00', end_color='FFFF00', fill_type='solid'
            )
    
    output.seek(0)
    return output

def to_excel_with_format_prod(tabla: pd.DataFrame) -> io.BytesIO:
    """
    Convierte una tabla de producción a Excel con formato
    
    Args:
        tabla: DataFrame a convertir
        
    Returns:
        BytesIO con el archivo Excel formateado
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        tabla.to_excel(writer, sheet_name='ProduccionDiaria')
    output.seek(0)
    return output

def to_excel_with_format_weekend(tabla: pd.DataFrame) -> io.BytesIO:
    """
    Convierte una tabla de fin de semana a Excel con formato
    
    Args:
        tabla: DataFrame a convertir
        
    Returns:
        BytesIO con el archivo Excel formateado
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        tabla.to_excel(writer, sheet_name='FinDeSemana')
    output.seek(0)
    return output

def to_excel_resumen(tabla: pd.DataFrame) -> io.BytesIO:
    """
    Convierte una tabla de resumen a Excel con formato
    
    Args:
        tabla: DataFrame a convertir
        
    Returns:
        BytesIO con el archivo Excel formateado
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        tabla.to_excel(writer, sheet_name='ResumenIngresos')
    output.seek(0)
    return output

def to_excel_matriz(tabla: pd.DataFrame) -> io.BytesIO:
    """
    Convierte una matriz de evolución de pendientes a Excel
    
    Args:
        tabla: DataFrame a convertir
        
    Returns:
        BytesIO con el archivo Excel
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        tabla.to_excel(writer, sheet_name='EvolucionPendientes')
    output.seek(0)
    return output 