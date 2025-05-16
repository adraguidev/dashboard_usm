import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path
import io
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import plotly.figure_factory as ff
from numpy import polyfit, arange
import numpy as np
import datetime
import pytz

# Configuración de la página
st.set_page_config(
    page_title="Dashboard de Procesos",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://www.example.com/help',
        'Report a bug': "https://www.example.com/bug",
        'About': "# Dashboard de Análisis de Procesos. Creado con Streamlit."
    }
)

# Título principal
st.title("Dashboard de Análisis de Procesos")

# Función para cargar los datos
@st.cache_data
def cargar_datos(archivo):
    return pd.read_excel(f"ARCHIVOS/{archivo}")

# Sidebar para selección de proceso
proceso = st.sidebar.selectbox(
    "Seleccione el Proceso",
    ["CCM", "PRR"]
)

# Mapeo de archivos
archivos = {
    "CCM": "consolidado_final_CCM_personal.xlsx",
    "PRR": "consolidado_final_PRR_personal.xlsx"
}

# Cargar datos según el proceso seleccionado
try:
    df = cargar_datos(archivos[proceso])
    
    # Crear pestañas para diferentes análisis
    tab1, tab2, tab3, tab5, tab6 = st.tabs([
        "Pendientes", 
        "Producción Diaria", 
        "Ingresos Diarios",
        "Proyección de Cierre",
        "Evolución Pendientes x Operador"
    ])
    
    with tab1:
        st.header("Pendientes CCM" if proceso == "CCM" else "Pendientes PRR")
        
        # Definir filtros según el proceso
        if proceso == "CCM":
            df_filtrado = df[
                (df['UltimaEtapa'] == 'EVALUACIÓN - I') &
                (df['EstadoPre'].isna()) &
                (df['EstadoTramite'] == 'PENDIENTE') &
                (df['EQUIPO'] != 'VULNERABLE')
            ]
        else:
            etapas_prr = [
                'ACTUALIZAR DATOS BENEFICIARIO - F',
                'ACTUALIZAR DATOS BENEFICIARIO - I',
                'ASOCIACION BENEFICIARIO - F',
                'ASOCIACION BENEFICIARIO - I',
                'CONFORMIDAD SUB-DIREC.INMGRA. - I',
                'PAGOS, FECHA Y NRO RD. - F',
                'PAGOS, FECHA Y NRO RD. - I',
                'RECEPCIÓN DINM - F'
            ]
            df_filtrado = df[
                (df['UltimaEtapa'].isin(etapas_prr)) &
                (df['EstadoPre'].isna()) &
                (df['EstadoTramite'] == 'PENDIENTE') &
                (df['EQUIPO'] != 'VULNERABLE')
            ]

        # Reemplazar nulos en OPERADOR por 'Sin asignar'
        df_filtrado['OPERADOR'] = df_filtrado['OPERADOR'].fillna('Sin asignar')

        # Crear tabla dinámica sin totales automáticos
        tabla = pd.pivot_table(
            df_filtrado,
            index='OPERADOR',
            columns='Anio',
            values='NumeroTramite',
            aggfunc='count',
            fill_value=0
        )
        # Calcular columna Total manualmente
        tabla['Total'] = tabla.sum(axis=1)
        # Excluir operadores específicos en CCM y Sin asignar en PRR
        if proceso == "CCM":
            operadores_excluir = ["MAURICIO ROMERO, HUGO", "Sin asignar"]
            tabla = tabla.drop(operadores_excluir, errors='ignore')
        elif proceso == "PRR":
            tabla = tabla.drop(["Sin asignar"], errors='ignore')
        # Ordenar por Total descendente
        tabla = tabla.sort_values(by=('Total'), ascending=False)
        # Recalcular la fila Total después de filtrar
        total_row = tabla.sum(axis=0)
        total_row.name = 'Total'
        tabla = pd.concat([tabla, pd.DataFrame([total_row])])
        # Mostrar tabla
        st.dataframe(tabla, use_container_width=True, height=500)
        # Mostrar el total de "Sin asignar" solo para los últimos 2 años
        anios = sorted(df_filtrado['Anio'].dropna().unique())
        ultimos_2_anios = anios[-2:] if len(anios) >= 2 else anios
        if proceso == "CCM":
            total_sin_asignar = df_filtrado[
                (df_filtrado['OPERADOR'] == 'Sin asignar') &
                (df_filtrado['Anio'].isin(ultimos_2_anios))
            ]['NumeroTramite'].count()
            st.metric("Sin asignar (últimos 2 años)", total_sin_asignar)
        elif proceso == "PRR":
            total_sin_asignar = df_filtrado[
                (df_filtrado['OPERADOR'] == 'Sin asignar') &
                (df_filtrado['Anio'].isin(ultimos_2_anios))
            ]['NumeroTramite'].count()
            st.metric("Sin asignar (últimos 2 años)", total_sin_asignar)
        
        # Botón para descargar Excel con formato
        def to_excel_with_format(tabla):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                tabla.to_excel(writer, sheet_name='Pendientes')
                ws = writer.sheets['Pendientes']
                # Resaltar fila Total
                total_row = ws.max_row
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=total_row, column=col).font = Font(bold=True)
                    ws.cell(row=total_row, column=col).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                # Resaltar columna Total
                total_col = ws.max_column
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=total_col).font = Font(bold=True)
                    ws.cell(row=row, column=total_col).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            output.seek(0)
            return output

        excel_data = to_excel_with_format(tabla)
        st.download_button(
            label="Descargar tabla en Excel",
            data=excel_data,
            file_name=f"pendientes_{proceso}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # --- Guardado automático del histórico de pendientes por operador y año ---
        # Quitar la fila 'Total' para el histórico
        tabla_historico = tabla.drop('Total', errors='ignore').copy()
        # Quitar la columna 'Total' para el histórico
        if 'Total' in tabla_historico.columns:
            tabla_historico = tabla_historico.drop(columns=['Total'])
        # Asegurar que 'OPERADOR' sea columna
        if 'OPERADOR' not in tabla_historico.columns:
            tabla_historico = tabla_historico.reset_index().rename(columns={tabla_historico.index.name or 'index': 'OPERADOR'})
        # Convertir a formato largo
        tabla_historico = tabla_historico.melt(
            id_vars=['OPERADOR'],
            var_name='Año',
            value_name='Pendientes'
        )
        # Agregar columnas de fecha local y proceso
        from datetime import datetime
        tz = pytz.timezone('America/Lima')
        fecha_hoy_local = datetime.now(tz).strftime('%Y-%m-%d')
        tabla_historico['Fecha'] = fecha_hoy_local
        tabla_historico['Proceso'] = proceso
        # Reordenar columnas
        tabla_historico = tabla_historico[['Fecha', 'Proceso', 'OPERADOR', 'Año', 'Pendientes']]
        # Normalizar claves para evitar duplicados por diferencias de formato
        tabla_historico['OPERADOR'] = tabla_historico['OPERADOR'].str.strip().str.upper()
        tabla_historico['Año'] = tabla_historico['Año'].astype(str)
        # Ruta del archivo histórico
        ruta_historico = 'ARCHIVOS/historico_pendientes_operador.csv'
        # Leer histórico existente si existe
        try:
            historico_existente = pd.read_csv(ruta_historico, dtype=str)
            historico_existente['OPERADOR'] = historico_existente['OPERADOR'].str.strip().str.upper()
            historico_existente['Año'] = historico_existente['Año'].astype(str)
        except FileNotFoundError:
            historico_existente = pd.DataFrame(columns=tabla_historico.columns)

        # --- Nueva lógica: solo actualizar si los valores de pendientes realmente cambiaron ---
        claves = ['Fecha', 'Proceso', 'OPERADOR', 'Año']
        if not historico_existente.empty:
            # Hacemos merge para comparar valores existentes y nuevos
            merge = tabla_historico.merge(
                historico_existente,
                on=claves,
                how='left',
                suffixes=('', '_old')
            )
            # Solo tomar los registros donde el valor de Pendientes es diferente o no existe
            tabla_historico_filtrada = merge[(merge['Pendientes_old'].isna()) | (merge['Pendientes'] != merge['Pendientes_old'])][tabla_historico.columns]
            # Eliminar del histórico existente los registros que vamos a actualizar
            historico_actualizado = historico_existente[~historico_existente.set_index(claves).index.isin(tabla_historico_filtrada.set_index(claves).index)]
            # Concatenar los nuevos registros (actualizados o nuevos)
            historico_actualizado = pd.concat([historico_actualizado, tabla_historico_filtrada], ignore_index=True)
        else:
            tabla_historico_filtrada = tabla_historico.copy()
            historico_actualizado = tabla_historico.copy()

        if not tabla_historico_filtrada.empty:
            historico_actualizado.to_csv(ruta_historico, index=False)

    with tab2:
        st.header("Producción Diaria")
        # Determinar nombres de columnas según proceso
        col_operador = 'OperadorPre'
        col_fecha = 'FechaPre'
        col_tramite = 'NumeroTramite'
        if col_operador not in df.columns:
            col_operador = 'OPERADOR'  # fallback por si acaso
        if col_fecha not in df.columns:
            col_fecha = 'FechaPre'  # fallback
        # Filtrar últimos 20 días
        if pd.api.types.is_datetime64_any_dtype(df[col_fecha]):
            fechas_ordenadas = df[col_fecha].dropna().sort_values().unique()
        else:
            df[col_fecha] = pd.to_datetime(df[col_fecha], errors='coerce')
            fechas_ordenadas = df[col_fecha].dropna().sort_values().unique()
        ultimos_20 = fechas_ordenadas[-20:]
        df_20dias = df[df[col_fecha].isin(ultimos_20)]
        # Crear tabla dinámica
        tabla_prod = pd.pivot_table(
            df_20dias,
            index=col_operador,
            columns=col_fecha,
            values=col_tramite,
            aggfunc='count',
            fill_value=0,
            margins=True,
            margins_name='Total'
        )
        # Excluir operadores específicos y aquellos con total menor a 5
        operadores_excluir = ["Aponte Sanchez, Paola Lita", "Lucero Martinez, Carlos Martin", "USUARIO DE AGENCIA DIGITAL"]
        if 'Total' in tabla_prod.index:
            tabla_filtrada = tabla_prod.drop(operadores_excluir, errors='ignore')
            # Excluir filas con total menor a 5 (excepto la fila Total)
            if tabla_filtrada.shape[0] > 1:
                tabla_filtrada = tabla_filtrada[ (tabla_filtrada['Total'] >= 5) | (tabla_filtrada.index == 'Total') ]
        else:
            tabla_filtrada = tabla_prod.drop(operadores_excluir, errors='ignore')
            tabla_filtrada = tabla_filtrada[tabla_filtrada['Total'] >= 5]

        # Recalcular la fila Total después de filtrar y ordenar (Producción Diaria)
        tabla_sin_total = tabla_filtrada.drop('Total', errors='ignore')
        total_row = tabla_sin_total.sum(numeric_only=True)
        total_row.name = 'Total'
        tabla_filtrada_corr = pd.concat([tabla_sin_total, pd.DataFrame([total_row])])
        # Formatear las fechas de las columnas al formato dd/mm/YYYY
        fechas_formateadas = [f.strftime('%d/%m/%Y') if not isinstance(f, str) and f != 'Total' else f for f in tabla_filtrada_corr.columns]
        tabla_filtrada_corr.columns = fechas_formateadas
        # Ordenar por Total descendente, dejando la fila 'Total' al final (Producción Diaria)
        if 'Total' in tabla_filtrada_corr.index:
            tabla_sin_total = tabla_filtrada_corr.drop('Total')
            tabla_sin_total = tabla_sin_total.sort_values(by='Total', ascending=False)
            tabla_filtrada_corr = pd.concat([tabla_sin_total, tabla_filtrada_corr.loc[['Total']]])
        else:
            tabla_filtrada_corr = tabla_filtrada_corr.sort_values(by='Total', ascending=False)

        st.dataframe(tabla_filtrada_corr, use_container_width=True, height=500)

        # Botón para descargar Excel con formato (Producción Diaria)
        def to_excel_with_format_prod(tabla):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                tabla.to_excel(writer, sheet_name='ProduccionDiaria')
                ws = writer.sheets['ProduccionDiaria']
                # Resaltar fila Total
                total_row = ws.max_row
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=total_row, column=col).font = Font(bold=True)
                    ws.cell(row=total_row, column=col).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                # Resaltar columna Total
                total_col = ws.max_column
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=total_col).font = Font(bold=True)
                    ws.cell(row=row, column=total_col).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            output.seek(0)
            return output
        excel_data_prod = to_excel_with_format_prod(tabla_filtrada_corr)
        st.download_button(
            label="Descargar tabla de Producción Diaria en Excel",
            data=excel_data_prod,
            file_name=f"produccion_diaria_{proceso}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # --- Tabla de fines de semana (últimas 5 semanas) ---
        st.subheader("Producción Fines de Semana (Últimas 5 semanas)")
        # Calcular el rango de fechas de las últimas 5 semanas
        fecha_max = df[col_fecha].max()
        fecha_min = fecha_max - pd.Timedelta(weeks=5)
        df_5sem = df[(df[col_fecha] >= fecha_min) & (df[col_fecha] <= fecha_max)]
        # Filtrar solo sábados (5) y domingos (6)
        df_5sem = df_5sem[df_5sem[col_fecha].dt.weekday.isin([5, 6])]
        # Crear tabla dinámica
        tabla_weekend = pd.pivot_table(
            df_5sem,
            index=col_operador,
            columns=col_fecha,
            values=col_tramite,
            aggfunc='count',
            fill_value=0,
            margins=True,
            margins_name='Total'
        )
        # Excluir operadores específicos y totales menores a 5
        if 'Total' in tabla_weekend.index:
            tabla_weekend_filtrada = tabla_weekend.drop(operadores_excluir, errors='ignore')
            if tabla_weekend_filtrada.shape[0] > 1:
                tabla_weekend_filtrada = tabla_weekend_filtrada[(tabla_weekend_filtrada['Total'] >= 5) | (tabla_weekend_filtrada.index == 'Total')]
        else:
            tabla_weekend_filtrada = tabla_weekend.drop(operadores_excluir, errors='ignore')
            tabla_weekend_filtrada = tabla_weekend_filtrada[tabla_weekend_filtrada['Total'] >= 5]
        # Recalcular la fila Total después de filtrar y ordenar (Fines de Semana)
        tabla_sin_total_w = tabla_weekend_filtrada.drop('Total', errors='ignore')
        total_row_w = tabla_sin_total_w.sum(numeric_only=True)
        total_row_w.name = 'Total'
        tabla_weekend_filtrada_corr = pd.concat([tabla_sin_total_w, pd.DataFrame([total_row_w])])
        # Formatear las fechas de las columnas al formato dd/mm/YYYY
        fechas_formateadas_w = [f.strftime('%d/%m/%Y') if not isinstance(f, str) and f != 'Total' else f for f in tabla_weekend_filtrada_corr.columns]
        tabla_weekend_filtrada_corr.columns = fechas_formateadas_w
        # Ordenar por Total descendente, dejando la fila 'Total' al final (Fines de Semana)
        if 'Total' in tabla_weekend_filtrada_corr.index:
            tabla_sin_total_w = tabla_weekend_filtrada_corr.drop('Total')
            tabla_sin_total_w = tabla_sin_total_w.sort_values(by='Total', ascending=False)
            tabla_weekend_filtrada_corr = pd.concat([tabla_sin_total_w, tabla_weekend_filtrada_corr.loc[['Total']]])
        else:
            tabla_weekend_filtrada_corr = tabla_weekend_filtrada_corr.sort_values(by='Total', ascending=False)

        st.dataframe(tabla_weekend_filtrada_corr, use_container_width=True, height=400)
        # Botón para descargar Excel con formato (Fines de Semana)
        def to_excel_with_format_weekend(tabla):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                tabla.to_excel(writer, sheet_name='FinesDeSemana')
                ws = writer.sheets['FinesDeSemana']
                # Resaltar fila Total
                total_row = ws.max_row
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=total_row, column=col).font = Font(bold=True)
                    ws.cell(row=total_row, column=col).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                # Resaltar columna Total
                total_col = ws.max_column
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=total_col).font = Font(bold=True)
                    ws.cell(row=row, column=total_col).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            output.seek(0)
            return output
        excel_data_weekend = to_excel_with_format_weekend(tabla_weekend_filtrada_corr)
        st.download_button(
            label="Descargar tabla de fines de semana en Excel",
            data=excel_data_weekend,
            file_name=f"produccion_fines_semana_{proceso}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # --- Tabla resumen por FechaPre ---
        st.subheader("Resumen Diario de Producción")
        # Operadores a excluir
        operadores_excluir_resumen = ["Aponte Sanchez, Paola Lita", "Lucero Martinez, Carlos Martin", "USUARIO DE AGENCIA DIGITAL"]
        # Filtrar el dataframe de los últimos 20 días
        df_resumen = df_20dias[~df_20dias[col_operador].isin(operadores_excluir_resumen)].copy()
        # Calcular el total por operador (en los últimos 20 días)
        totales_operador = df_resumen.groupby(col_operador)[col_tramite].count()
        operadores_validos = totales_operador[totales_operador >= 5].index
        df_resumen = df_resumen[df_resumen[col_operador].isin(operadores_validos)]
        # Calcular cantidad de operadores y total de trámites por FechaPre
        resumen = df_resumen.groupby(col_fecha).agg(
            cantidad_operadores=(col_operador, lambda x: x.nunique()),
            total_trabajados=(col_tramite, 'count')
        )
        resumen = resumen.sort_index()
        resumen['promedio_por_operador'] = resumen['total_trabajados'] / resumen['cantidad_operadores']
        # Formatear fechas
        resumen.index = [f.strftime('%d/%m/%Y') if not isinstance(f, str) else f for f in resumen.index]
        st.dataframe(resumen, use_container_width=True, height=400)
        # Botón para descargar Excel
        def to_excel_resumen(tabla):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                tabla.to_excel(writer, sheet_name='Resumen')
            output.seek(0)
            return output
        excel_data_resumen = to_excel_resumen(resumen)
        st.download_button(
            label="Descargar resumen diario en Excel",
            data=excel_data_resumen,
            file_name=f"resumen_diario_{proceso}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # --- Gráficos representativos ---
        st.subheader("Gráfica: Promedio Diario por Operador (Lunes a Viernes)")
        resumen_graf_habiles = resumen.copy()
        resumen_graf_habiles.index = pd.to_datetime(resumen_graf_habiles.index, format='%d/%m/%Y')
        dias_habiles = resumen_graf_habiles.index.weekday < 5
        resumen_graf_habiles = resumen_graf_habiles[dias_habiles]
        fig_habiles = go.Figure()
        fig_habiles.add_trace(go.Scatter(
            x=resumen_graf_habiles.index,
            y=resumen_graf_habiles['promedio_por_operador'],
            mode='lines+markers+text',
            name='Promedio por Operador (L-V)',
            text=[f"{v:.1f}" for v in resumen_graf_habiles['promedio_por_operador']],
            textposition="top center"
        ))
        # Línea de tendencia
        x_numeric = arange(len(resumen_graf_habiles.index))
        y = resumen_graf_habiles['promedio_por_operador'].values
        if len(x_numeric) > 1:
            z = polyfit(x_numeric, y, 1)
            tendencia = z[0] * x_numeric + z[1]
            fig_habiles.add_trace(go.Scatter(
                x=resumen_graf_habiles.index,
                y=tendencia,
                mode='lines',
                name='Tendencia',
                line=dict(dash='dash', color='orange')
            ))
        fig_habiles.update_layout(
            title='Promedio Diario de Trámites por Operador (Lunes a Viernes)',
            xaxis_title='Fecha',
            yaxis_title='Promedio por Operador',
            legend_title='Métrica',
            hovermode='x unified'
        )
        st.plotly_chart(fig_habiles, use_container_width=True)

        st.subheader("Gráfica: Promedio Diario por Operador (Fines de Semana)")
        resumen_graf_fds = resumen.copy()
        resumen_graf_fds.index = pd.to_datetime(resumen_graf_fds.index, format='%d/%m/%Y')
        dias_fds = resumen_graf_fds.index.weekday >= 5
        resumen_graf_fds = resumen_graf_fds[dias_fds]
        fig_fds = go.Figure()
        fig_fds.add_trace(go.Scatter(
            x=resumen_graf_fds.index,
            y=resumen_graf_fds['promedio_por_operador'],
            mode='lines+markers+text',
            name='Promedio por Operador (S-D)',
            text=[f"{v:.1f}" for v in resumen_graf_fds['promedio_por_operador']],
            textposition="top center"
        ))
        # Línea de tendencia
        x_numeric_fds = arange(len(resumen_graf_fds.index))
        y_fds = resumen_graf_fds['promedio_por_operador'].values
        if len(x_numeric_fds) > 1:
            z_fds = polyfit(x_numeric_fds, y_fds, 1)
            tendencia_fds = z_fds[0] * x_numeric_fds + z_fds[1]
            fig_fds.add_trace(go.Scatter(
                x=resumen_graf_fds.index,
                y=tendencia_fds,
                mode='lines',
                name='Tendencia',
                line=dict(dash='dash', color='orange')
            ))
        fig_fds.update_layout(
            title='Promedio Diario de Trámites por Operador (Fines de Semana)',
            xaxis_title='Fecha',
            yaxis_title='Promedio por Operador',
            legend_title='Métrica',
            hovermode='x unified'
        )
        st.plotly_chart(fig_fds, use_container_width=True)

        st.subheader("Gráfica: Total de Trámites Diarios")
        resumen_graf_total = resumen.copy()
        resumen_graf_total.index = pd.to_datetime(resumen_graf_total.index, format='%d/%m/%Y')
        fig_total = go.Figure()
        fig_total.add_trace(go.Scatter(
            x=resumen_graf_total.index,
            y=resumen_graf_total['total_trabajados'],
            mode='lines+markers+text',
            name='Total de Trámites',
            text=[str(v) for v in resumen_graf_total['total_trabajados']],
            textposition="top center"
        ))
        # Línea de tendencia
        x_numeric_total = arange(len(resumen_graf_total.index))
        y_total = resumen_graf_total['total_trabajados'].values
        if len(x_numeric_total) > 1:
            z_total = polyfit(x_numeric_total, y_total, 1)
            tendencia_total = z_total[0] * x_numeric_total + z_total[1]
            fig_total.add_trace(go.Scatter(
                x=resumen_graf_total.index,
                y=tendencia_total,
                mode='lines',
                name='Tendencia',
                line=dict(dash='dash', color='orange')
            ))
        fig_total.update_layout(
            title='Total de Trámites Diarios',
            xaxis_title='Fecha',
            yaxis_title='Total de Trámites',
            legend_title='Métrica',
            hovermode='x unified'
        )
        st.plotly_chart(fig_total, use_container_width=True)

    with tab3:
        st.header("Ingreso de Expedientes")
        # Determinar columna de fecha (corregido a FechaExpendiente)
        col_fecha_ing = 'FechaExpendiente'
        col_tramite_ing = 'NumeroTramite'
        if col_fecha_ing not in df.columns:
            st.warning("No se encontró la columna FechaExpendiente en los datos.")
        else:
            # Asegurar tipo datetime
            df[col_fecha_ing] = pd.to_datetime(df[col_fecha_ing], errors='coerce')
            # Filtrar últimos 60 días
            fecha_max = df[col_fecha_ing].max()
            fecha_min = fecha_max - pd.Timedelta(days=60)
            df_60dias = df[(df[col_fecha_ing] >= fecha_min) & (df[col_fecha_ing] <= fecha_max)]
            # Agrupar por fecha y contar NumeroTramite
            ingresos_diarios = df_60dias.groupby(col_fecha_ing)[col_tramite_ing].count().reset_index()
            ingresos_diarios = ingresos_diarios.sort_values(col_fecha_ing)
            # Gráfico
            fig = go.Figure()
            # Línea y puntos
            fig.add_trace(go.Scatter(
                x=ingresos_diarios[col_fecha_ing],
                y=ingresos_diarios[col_tramite_ing],
                mode='lines+markers+text',
                name='NumeroTramite',
                text=[str(v) for v in ingresos_diarios[col_tramite_ing]],
                textposition="top center",
                line=dict(color='royalblue'),
                fill='tozeroy',
                fillcolor='rgba(65,105,225,0.1)'
            ))
            # Línea de tendencia
            x_numeric = arange(len(ingresos_diarios))
            y_vals = ingresos_diarios[col_tramite_ing].values # Renombrado para evitar conflicto
            if len(x_numeric) > 1:
                z = polyfit(x_numeric, y_vals, 1)
                tendencia = z[0] * x_numeric + z[1]
                fig.add_trace(go.Scatter(
                    x=ingresos_diarios[col_fecha_ing],
                    y=tendencia,
                    mode='lines',
                    name='Tendencia',
                    line=dict(dash='dash', color='red', width=3)
                ))
            # Formato de fechas en eje X
            fig.update_xaxes(
                tickformat="%d %b",
                tickangle=0
            )
            fig.update_layout(
                title='',
                xaxis_title='',
                yaxis_title='',
                legend_title='',
                hovermode='x unified',
                margin=dict(l=20, r=20, t=40, b=20)
            )
            st.plotly_chart(fig, use_container_width=True)

            # Tabla de los últimos 15 días
            st.write("#### Ingresos diarios - últimos 15 días")
            tabla_15 = ingresos_diarios.tail(15).copy()
            tabla_15['FechaExpendiente'] = tabla_15['FechaExpendiente'].dt.strftime('%d/%m/%Y')
            tabla_15 = tabla_15.rename(columns={'FechaExpendiente': 'Fecha', 'NumeroTramite': 'Ingresos'})
            st.dataframe(tabla_15, use_container_width=True)
            # Gráfico de promedio semanal
            st.write("#### Promedio semanal de ingresos diarios")
            if col_fecha_ing not in df.columns:
                st.warning("No se encontró la columna FechaExpendiente en los datos.") 
            else:
                df_sem = df.copy()
                df_sem[col_fecha_ing] = pd.to_datetime(df_sem[col_fecha_ing], errors='coerce')
                fecha_max_sem = df_sem[col_fecha_ing].max() 
                fecha_min_sem = fecha_max_sem - pd.Timedelta(days=365) 
                df_sem = df_sem[(df_sem[col_fecha_ing] >= fecha_min_sem) & (df_sem[col_fecha_ing] <= fecha_max_sem)]
                df_sem['Semana'] = df_sem[col_fecha_ing].dt.to_period('W').dt.start_time
                ingresos_diarios_semanal = df_sem.groupby('Semana')[col_tramite_ing].count().reset_index()
                ingresos_diarios_semanal = ingresos_diarios_semanal.rename(columns={col_tramite_ing: 'Total ingresos'})
                ingresos_diarios_semanal['Promedio semanal'] = ingresos_diarios_semanal['Total ingresos'] / 7
                ingresos_diarios_semanal['Fecha'] = ingresos_diarios_semanal['Semana'].dt.strftime('%d/%m/%Y')
                ingresos_diarios_semanal['Rango de fechas'] = ingresos_diarios_semanal['Semana'].dt.strftime('%d/%m/%Y') + ' - ' + (ingresos_diarios_semanal['Semana'] + pd.Timedelta(days=6)).dt.strftime('%d/%m/%Y')
                semana_actual = pd.Timestamp.today().to_period('W').start_time
                ingresos_diarios_semanal['Es semana actual'] = ingresos_diarios_semanal['Semana'] == semana_actual
                fig_sem = px.line(
                    ingresos_diarios_semanal,
                    x='Fecha',
                    y='Promedio semanal',
                    title='Promedio semanal de ingresos diarios (último año)',
                    labels={'Fecha': 'Fecha', 'Promedio semanal': 'Promedio semanal de ingresos'},
                    hover_data={'Rango de fechas': True}
                )
                fig_sem.update_traces(mode='lines+markers', marker=dict(color=ingresos_diarios_semanal['Es semana actual'].map({True: 'red', False: 'blue'})))
                anio_actual = pd.Timestamp.today().year
                mask_anio = ingresos_diarios_semanal['Semana'].dt.year == anio_actual
                sem_actual = ingresos_diarios_semanal[mask_anio].reset_index(drop=True)
                if len(sem_actual) > 1:
                    x_numeric_sem = np.arange(len(sem_actual)) 
                    y_vals_sem = sem_actual['Promedio semanal'].values 
                    z_sem = np.polyfit(x_numeric_sem, y_vals_sem, 1) 
                    tendencia_sem = z_sem[0] * x_numeric_sem + z_sem[1] 
                    fig_sem.add_scatter(
                        x=sem_actual['Fecha'],
                        y=tendencia_sem,
                        mode='lines',
                        name='Tendencia año en curso',
                        line=dict(dash='dash', color='orange')
                    )
                fig_sem.update_xaxes(tickangle=45)
                st.plotly_chart(fig_sem, use_container_width=True)
            st.write("""**¿Qué muestra este gráfico?**
- Permite ver si el tiempo promedio para pretrabajar un expediente ha mejorado o empeorado a lo largo del año.
- Una tendencia descendente indica mayor eficiencia; una ascendente, posibles cuellos de botella o sobrecarga.""")

    with tab5:
        st.header("Proyección de Cierre y Equilibrio")
        
        # --- 1. Cálculos Base ---
        # a. Pendientes actuales (del mismo filtro que tab1 y usado para el default de personal_activo_input)
        if proceso == "CCM":
            df_pend_calc = df[
                (df['UltimaEtapa'] == 'EVALUACIÓN - I') &
                (df['EstadoPre'].isna()) &
                (df['EstadoTramite'] == 'PENDIENTE') &
                (df['EQUIPO'] != 'VULNERABLE')
            ]
        else:
            etapas_prr = [
                'ACTUALIZAR DATOS BENEFICIARIO - F', 'ACTUALIZAR DATOS BENEFICIARIO - I',
                'ASOCIACION BENEFICIARIO - F', 'ASOCIACION BENEFICIARIO - I',
                'CONFORMIDAD SUB-DIREC.INMGRA. - I', 'PAGOS, FECHA Y NRO RD. - F',
                'PAGOS, FECHA Y NRO RD. - I', 'RECEPCIÓN DINM - F'
            ]
            df_pend_calc = df[
                (df['UltimaEtapa'].isin(etapas_prr)) &
                (df['EstadoPre'].isna()) &
                (df['EstadoTramite'] == 'PENDIENTE') &
                (df['EQUIPO'] != 'VULNERABLE')
            ]
        pendientes_actuales_totales = len(df_pend_calc)
        pendientes_sin_asignar_actuales = df_pend_calc['OPERADOR'].isna().sum()
        pendientes_asignados_actuales = pendientes_actuales_totales - pendientes_sin_asignar_actuales

        # b. Ingresos diarios promedio (últimos 60 días)
        df['FechaExpendiente'] = pd.to_datetime(df['FechaExpendiente'], errors='coerce')
        fecha_max_ingresos = df['FechaExpendiente'].max()
        fecha_min_ingresos = fecha_max_ingresos - pd.Timedelta(days=60)
        ingresos_ultimos_60d = df[(df['FechaExpendiente'] >= fecha_min_ingresos) & (df['FechaExpendiente'] <= fecha_max_ingresos)]
        ingresos_diarios_promedio = ingresos_ultimos_60d.groupby(df['FechaExpendiente'].dt.date)['NumeroTramite'].count().mean()
        if pd.isna(ingresos_diarios_promedio): ingresos_diarios_promedio = 0

        # c. Productividad individual promedio (cierres diarios por persona, últimos 20 días)
        # (Misma lógica que antes para 'promedio_por_operador')
        col_operador_prod = 'OperadorPre'
        col_fecha_prod = 'FechaPre'
        col_tramite_prod = 'NumeroTramite'
        if col_operador_prod not in df.columns: col_operador_prod = 'OPERADOR'
        if col_fecha_prod not in df.columns: col_fecha_prod = 'FechaPre'
        
        if pd.api.types.is_datetime64_any_dtype(df[col_fecha_prod]):
            fechas_ordenadas_prod = df[col_fecha_prod].dropna().sort_values().unique()
        else:
            df[col_fecha_prod] = pd.to_datetime(df[col_fecha_prod], errors='coerce')
            fechas_ordenadas_prod = df[col_fecha_prod].dropna().sort_values().unique()
        
        ultimos_20_dias_prod = fechas_ordenadas_prod[-20:]
        df_20dias_prod = df[df[col_fecha_prod].isin(ultimos_20_dias_prod)]
        
        operadores_excluir_prod = [
            "Aponte Sanchez, Paola Lita", "Lucero Martinez, Carlos Martin",
            "USUARIO DE AGENCIA DIGITAL", "MAURICIO ROMERO, HUGO", "Sin asignar"
        ]
        df_20dias_prod = df_20dias_prod[~df_20dias_prod[col_operador_prod].isin(operadores_excluir_prod)]
        
        totales_operador_prod = df_20dias_prod.groupby(col_operador_prod)[col_tramite_prod].count()
        operadores_validos_prod = totales_operador_prod[totales_operador_prod >= 5].index
        df_20dias_prod = df_20dias_prod[df_20dias_prod[col_operador_prod].isin(operadores_validos_prod)]
        
        resumen_prod_diaria = df_20dias_prod.groupby(df[col_fecha_prod].dt.date).agg(
            cantidad_operadores=(col_operador_prod, lambda x: x.nunique()),
            total_trabajados=(col_tramite_prod, 'count')
        )
        if not resumen_prod_diaria.empty and 'cantidad_operadores' in resumen_prod_diaria and 'total_trabajados' in resumen_prod_diaria:
            resumen_prod_diaria = resumen_prod_diaria[resumen_prod_diaria['cantidad_operadores'] > 0] # Evitar división por cero
            resumen_prod_diaria['promedio_por_operador'] = resumen_prod_diaria['total_trabajados'] / resumen_prod_diaria['cantidad_operadores']
            productividad_individual_promedio = resumen_prod_diaria['promedio_por_operador'].mean()
        else:
            productividad_individual_promedio = 0
        if pd.isna(productividad_individual_promedio) or productividad_individual_promedio == 0 : 
            # Fallback si no hay datos de producción o es cero, para evitar divisiones por cero más adelante
            # Podrías ajustar este valor o manejarlo de otra forma si es necesario.
            productividad_individual_promedio = 1 # Asumir al menos 1 para evitar errores, o manejar con mensajes
            st.warning("No se pudo calcular la productividad individual promedio o es cero. Se usará un valor de 1 para cálculos. Revise los datos de producción.")


        # d. Personal activo por defecto (operadores con >= 5 pendientes)
        operadores_con_pendientes = df_pend_calc.groupby('OPERADOR').size()
        operadores_con_min_pendientes = operadores_con_pendientes[operadores_con_pendientes >= 5]
        num_operadores_activos_defecto = len(operadores_con_min_pendientes)
        if num_operadores_activos_defecto == 0: num_operadores_activos_defecto = 1 # Evitar que sea 0 por defecto

        # --- 2. Input del Usuario ---
        st.write("### Configure la Simulación")
        personal_simulacion = st.number_input(
            "Cantidad de personal activo para la simulación:",
            min_value=1, max_value=100, value=num_operadores_activos_defecto, step=1
        )

        # --- 3. Cálculos de Proyección ---
        cierres_estimados_diarios_simulacion = productividad_individual_promedio * personal_simulacion
        balance_diario_proyectado = cierres_estimados_diarios_simulacion - ingresos_diarios_promedio

        # --- 4. Presentación del Resumen ---
        st.write("### Resumen de Proyección")
        
        data_resumen = {
            'Métrica': [
                "**SITUACIÓN ACTUAL**",
                "Pendientes Totales Actuales",
                "Pendientes Asignados",
                "Pendientes Sin Asignar",
                "Ingresos Diarios Promedio (últimos 60 días)",
                "Productividad Individual Promedio (cierres/persona/día, últimos 20 días)",
                "**SIMULACIÓN CON PERSONAL CONFIGURADO**",
                "Personal Activo en Simulación",
                "Cierres Diarios Estimados (total equipo)",
                "Balance Diario Proyectado (Cierres Estimados - Ingresos Promedio)",
                "Proyección de Agotamiento de Pendientes Actuales",
                "**ANÁLISIS DE EQUILIBRIO DE FLUJO (Ingresos = Cierres)**",
                "Personal Necesario para Equilibrio de Flujo (aprox.)"
            ],
            'Valor': [
                "", # Separador
                f"{pendientes_actuales_totales}",
                f"{pendientes_asignados_actuales}",
                f"{pendientes_sin_asignar_actuales}",
                f"{ingresos_diarios_promedio:.2f}",
                f"{productividad_individual_promedio:.2f}",
                "", # Separador
                f"{personal_simulacion}",
                f"{cierres_estimados_diarios_simulacion:.2f}",
                "", # Llenado dinámico abajo
                "", # Llenado dinámico abajo
                "", # Separador
                ""  # Llenado dinámico abajo
            ]
        }

        # Llenado dinámico de Balance Diario
        if balance_diario_proyectado > 0:
            data_resumen['Valor'][9] = f"Superávit de {balance_diario_proyectado:.2f} expedientes/día. (Pendientes tienden a disminuir)"
        elif balance_diario_proyectado < 0:
            data_resumen['Valor'][9] = f"Déficit de {abs(balance_diario_proyectado):.2f} expedientes/día. (Pendientes tienden a aumentar)"
        else:
            data_resumen['Valor'][9] = "Equilibrio: 0 expedientes/día. (Pendientes tienden a mantenerse estables)"

        # Llenado dinámico de Días para Cero Pendientes
        if balance_diario_proyectado > 0:
            dias_para_cero_pendientes = pendientes_actuales_totales / balance_diario_proyectado
            data_resumen['Valor'][10] = f"{dias_para_cero_pendientes:.1f} días (si el ritmo se mantiene)"
        else:
            data_resumen['Valor'][10] = "No se agotarán los pendientes actuales con este ritmo."
            dias_para_cero_pendientes = float('inf') # Para el gráfico

        # Llenado dinámico de Personal para Equilibrio
        if productividad_individual_promedio > 0:
            personal_eq_flujo = np.ceil(ingresos_diarios_promedio / productividad_individual_promedio)
            data_resumen['Valor'][12] = f"{int(personal_eq_flujo)} personas"
        else:
            data_resumen['Valor'][12] = "No calculable (productividad individual es cero o no disponible)"

        resumen_df = pd.DataFrame(data_resumen)
        st.dataframe(resumen_df, use_container_width=True, hide_index=True)
        
        st.write("""
        **Interpretación del Resumen:**
        - **Situación Actual:** Muestra el estado actual de pendientes, ingresos y la capacidad de cierre individual.
        - **Simulación:** Proyecta el impacto del personal configurado. El "Balance Diario" es clave:
            - *Superávit*: Se cierran más expedientes de los que ingresan diariamente.
            - *Déficit*: Ingresan más expedientes de los que se cierran diariamente.
            - *Equilibrio*: Los cierres diarios igualan a los ingresos.
        - **Análisis de Equilibrio de Flujo:** Indica cuántas personas se necesitarían para que la cantidad de expedientes cerrados por día iguale la cantidad de expedientes que ingresan por día. Este es el punto donde el *stock* de pendientes dejaría de crecer.
        """)

        # --- 5. Gráfico de Proyección ---
        st.write("### Gráfico de Proyección de Pendientes")
        
        max_dias_grafico = 180 
        if balance_diario_proyectado > 0 and dias_para_cero_pendientes != float('inf'):
             # Si se proyecta cierre, mostrar un poco más allá del cierre o hasta 180 días
            max_dias_grafico = min(max_dias_grafico, int(dias_para_cero_pendientes) + 30)
        
        dias_proy = list(range(0, max_dias_grafico + 1))
        
        # Pendientes proyectados: stock_inicial - (balance_diario * dias)
        # balance_diario = cierres_estimados - ingresos_promedio
        # pendientes_proy = pendientes_actuales - (cierres_estimados - ingresos_promedio) * dias
        # pendientes_proy = pendientes_actuales - cierres_estimados * dias + ingresos_promedio * dias
        pendientes_proyectados_graf = [max(0, pendientes_actuales_totales - balance_diario_proyectado * d) for d in dias_proy]
        
        # Para el gráfico, también podemos mostrar las líneas de ingresos y cierres acumulados si es útil,
        # pero la línea de pendientes proyectados ya los considera.
        # Mostremos solo la evolución del stock de pendientes.

        fig_proy = go.Figure()
        fig_proy.add_trace(go.Scatter(
            x=dias_proy, 
            y=pendientes_proyectados_graf, 
            mode='lines+markers', 
            name='Pendientes Proyectados'
        ))

        fig_proy.update_layout(
            title='Evolución Estimada del Total de Pendientes',
            xaxis_title='Días desde Hoy',
            yaxis_title='Cantidad de Pendientes',
            hovermode='x unified'
        )
        
        # Añadir línea de ingresos y cierres estimados si se desea para comparación
        # (opcional, ya que el balance está en la línea de pendientes)
        # fig_proy.add_trace(go.Scatter(x=dias_proy, y=[ingresos_diarios_promedio * d for d in dias_proy], mode='lines', name='Ingresos Acumulados Estimados', line=dict(dash='dot')))
        # fig_proy.add_trace(go.Scatter(x=dias_proy, y=[cierres_estimados_diarios_simulacion * d for d in dias_proy], mode='lines', name='Cierres Acumulados Estimados', line=dict(dash='dot')))

        st.plotly_chart(fig_proy, use_container_width=True)
        
        st.write(f"""
        **Interpretación del Gráfico:**
        - El gráfico muestra cómo se proyecta que evolucione el **stock total de pendientes** día a día, considerando:
            - El total de pendientes actuales.
            - Los ingresos diarios promedio estimados ({ingresos_diarios_promedio:.2f}).
            - Los cierres diarios estimados con el personal configurado ({cierres_estimados_diarios_simulacion:.2f}).
        - **Si la línea desciende:** Los cierres superan a los ingresos. El punto donde cruza el eje X (cero pendientes) es la estimación de días para agotar los pendientes actuales.
        - **Si la línea asciende:** Los ingresos superan a los cierres. Los pendientes aumentarán.
        - **Si la línea es horizontal:** Los ingresos igualan a los cierres. El stock de pendientes se mantendría estable.
        Este gráfico asume que los promedios de ingresos y la productividad individual se mantienen constantes durante el período de proyección.
        """)

    # --- Nueva pestaña: Evolución Pendientes x Operador ---
    with tab6:
        st.header("Evolución de Pendientes por Operador")
        def cargar_historico():
            ruta_historico = 'ARCHIVOS/historico_pendientes_operador.csv'
            try:
                return pd.read_csv(ruta_historico, dtype={'Año': str})
            except FileNotFoundError:
                return pd.DataFrame(columns=['Fecha', 'Proceso', 'OPERADOR', 'Año', 'Pendientes'])
        recargar = st.button("Recargar solo histórico")
        historico = cargar_historico()
        # Agrupar años menores a 2024 como 'ANTIGUOS' para el filtro
        historico['Año'] = historico['Año'].apply(lambda x: 'ANTIGUOS' if (x.isdigit() and int(x) < 2024) or x == 'ANTIGUOS' else x)
        # Filtros
        # Solo mostrar el proceso seleccionado en el sidebar
        proceso_sel = proceso
        anios_disp = historico[historico['Proceso'] == proceso_sel]['Año'].unique().tolist()
        anios_disp = sorted(set(anios_disp), reverse=True, key=lambda x: (x != 'ANTIGUOS', x))
        anios_sel = st.multiselect("Año(s)", options=['Todos'] + anios_disp, default=['Todos'])
        if 'Todos' in anios_sel or not anios_sel:
            # Mostrar solo fechas que existen en todos los años
            anios_validos = [a for a in anios_disp if a != 'Todos']
            fechas_por_anio = [set(historico[(historico['Proceso'] == proceso_sel) & (historico['Año'] == anio)]['Fecha'].unique()) for anio in anios_validos]
            if fechas_por_anio:
                fechas_comunes = set.intersection(*fechas_por_anio)
            else:
                fechas_comunes = set()
            df_filtro = historico[(historico['Proceso'] == proceso_sel) & (historico['Fecha'].isin(fechas_comunes))].copy()
        elif len(anios_sel) > 1:
            # Mostrar solo fechas que existen en todos los años seleccionados
            fechas_por_anio = [set(historico[(historico['Proceso'] == proceso_sel) & (historico['Año'] == anio)]['Fecha'].unique()) for anio in anios_sel]
            if fechas_por_anio:
                fechas_comunes = set.intersection(*fechas_por_anio)
            else:
                fechas_comunes = set()
            df_filtro = historico[(historico['Proceso'] == proceso_sel) & (historico['Año'].isin(anios_sel)) & (historico['Fecha'].isin(fechas_comunes))].copy()
        else:
            df_filtro = historico[(historico['Proceso'] == proceso_sel) & (historico['Año'].isin(anios_sel))].copy()
        # Pivotear: filas=OPERADOR, columnas=Fecha, valores=Pendientes
        tabla_matriz = df_filtro.pivot_table(
            index='OPERADOR',
            columns='Fecha',
            values='Pendientes',
            aggfunc='sum',
            fill_value=0
        )
        # Ordenar columnas por fecha
        tabla_matriz = tabla_matriz.reindex(sorted(tabla_matriz.columns), axis=1)
        # Ordenar filas de mayor a menor según la última fecha disponible
        if len(tabla_matriz.columns) > 0:
            ultima_fecha = tabla_matriz.columns[-1]
            tabla_matriz = tabla_matriz.sort_values(by=ultima_fecha, ascending=False)
        # Agregar fila TOTAL
        total_row = tabla_matriz.sum(axis=0)
        total_row.name = 'TOTAL'
        tabla_matriz = pd.concat([tabla_matriz, pd.DataFrame([total_row])])
        st.dataframe(tabla_matriz, use_container_width=True, height=500)
        # Botón para descargar Excel
        def to_excel_matriz(tabla):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                tabla.to_excel(writer, sheet_name='EvolucionPendientes')
            output.seek(0)
            return output
        excel_data_matriz = to_excel_matriz(tabla_matriz)
        st.download_button(
            label="Descargar matriz en Excel",
            data=excel_data_matriz,
            file_name=f"evolucion_pendientes_{proceso_sel}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # Gráfico de totales por fecha
        st.subheader("Evolución de los totales de pendientes")
        totales = total_row.drop('TOTAL', errors='ignore')
        totales = totales.astype(int)
        fig_totales = go.Figure()
        # Línea principal con etiquetas
        fig_totales.add_trace(go.Scatter(
            x=totales.index,
            y=totales.values,
            mode='lines+markers+text',
            name='Total de pendientes',
            text=[str(v) for v in totales.values],
            textposition="top center"
        ))
        # Línea de tendencia sobre los últimos 7 días
        if len(totales) >= 2:
            ultimos_7 = totales[-7:]
            x_numeric = list(range(len(ultimos_7)))
            y = ultimos_7.values
            if len(x_numeric) > 1:
                z = np.polyfit(x_numeric, y, 1)
                tendencia = z[0] * np.array(x_numeric) + z[1]
                fig_totales.add_trace(go.Scatter(
                    x=ultimos_7.index,
                    y=tendencia,
                    mode='lines',
                    name='Tendencia (últimos 7 días)',
                    line=dict(dash='dash', color='orange')
                ))
        fig_totales.update_layout(
            title='Total de pendientes por fecha',
            xaxis_title='Fecha',
            yaxis_title='Total de pendientes',
            hovermode='x unified'
        )
        st.plotly_chart(fig_totales, use_container_width=True)

        # --- Ranking de evolución de pendientes por operador ---
        st.subheader("Ranking de evolución de pendientes por operador")
        # Selector de periodo
        opciones_periodo = [7, 15, 30, 'Todo el periodo']
        periodo_sel = st.selectbox("Periodo de análisis (días)", opciones_periodo, index=1)
        
        # Solo operadores (sin TOTAL)
        tabla_operadores = tabla_matriz.drop('TOTAL', errors='ignore')
        
        # Definir siempre tabla_operadores_periodo
        if len(tabla_operadores.columns) > 0:
            if periodo_sel == 'Todo el periodo':
                cols_periodo = tabla_operadores.columns
            else:
                cols_periodo = tabla_operadores.columns[-periodo_sel:]
            
            # Asegurarnos que OPERADOR sea una columna
            tabla_operadores = tabla_operadores.reset_index()
            if tabla_operadores.columns[0] != 'OPERADOR':
                tabla_operadores = tabla_operadores.rename(columns={tabla_operadores.columns[0]: 'OPERADOR'})
            
            tabla_operadores_periodo = tabla_operadores[['OPERADOR'] + cols_periodo.tolist()]
            tabla_operadores_filtrada = tabla_operadores_periodo[tabla_operadores_periodo[cols_periodo[-1]] >= 5]
            
            # Realizar el melt con la estructura correcta
            pendientes_long = tabla_operadores_filtrada.melt(
                id_vars=['OPERADOR'],
                value_vars=cols_periodo,
                var_name='Fecha',
                value_name='Pendientes'
            )
            
            # Normalizar operador y fecha
            pendientes_long['OPERADOR_NORM'] = pendientes_long['OPERADOR'].str.strip().str.upper()
            pendientes_long['Fecha'] = pd.to_datetime(pendientes_long['Fecha'], errors='coerce').dt.strftime('%Y-%m-%d')
            
            # Obtener datos de producción diaria
            col_operador = 'OperadorPre'
            col_fecha = 'FechaPre'
            col_tramite = 'NumeroTramite'
            
            if col_operador not in df.columns:
                col_operador = 'OPERADOR'
            if col_fecha not in df.columns:
                col_fecha = 'FechaPre'
            
            # Convertir fechas a datetime si no lo son
            if not pd.api.types.is_datetime64_any_dtype(df[col_fecha]):
                df[col_fecha] = pd.to_datetime(df[col_fecha], errors='coerce')
            
            # Filtrar por el periodo seleccionado
            fecha_min = pd.to_datetime(cols_periodo[0])
            fecha_max = pd.to_datetime(cols_periodo[-1])
            df_prod = df[(df[col_fecha] >= fecha_min) & (df[col_fecha] <= fecha_max)]
            
            # Calcular producción diaria por operador SOLO para el periodo seleccionado
            fechas_periodo = set([str(f) for f in cols_periodo])
            prod_diaria = df_prod.groupby([col_operador, col_fecha])[col_tramite].count().reset_index()
            prod_diaria[col_operador] = prod_diaria[col_operador].str.strip().str.upper()
            prod_diaria[col_fecha] = prod_diaria[col_fecha].dt.strftime('%Y-%m-%d')
            # Filtrar solo fechas del periodo seleccionado
            prod_diaria = prod_diaria[prod_diaria[col_fecha].isin(fechas_periodo)]
            # Calcular producción promedio y días de producción solo en el periodo
            prod_promedio = prod_diaria.groupby(col_operador)[col_tramite].agg(['mean', 'count']).reset_index()
            prod_promedio.columns = [col_operador, 'Produccion_Promedio', 'Dias_Produccion']
            prod_promedio[col_operador] = prod_promedio[col_operador].str.strip().str.upper()
            
            # Calcular métricas de evolución
            if len(pendientes_long) > 0:
                # Agrupar por operador y calcular métricas de pendientes
                evolucion = pendientes_long.groupby('OPERADOR_NORM').agg({
                    'Pendientes': ['first', 'last', 'count']
                }).reset_index()
                
                evolucion.columns = ['OPERADOR_NORM', 'Pendientes_Inicial', 'Pendientes_Final', 'Dias']
                evolucion['Cambio'] = evolucion['Pendientes_Final'] - evolucion['Pendientes_Inicial']
                
                # Corregir cálculo de Cambio_Porcentual para evitar inf
                evolucion['Cambio_Porcentual'] = evolucion.apply(
                    lambda row: (
                        (row['Cambio'] / row['Pendientes_Inicial'] * 100)
                        if row['Pendientes_Inicial'] != 0
                        else (np.sign(row['Cambio']) * 100.0 if row['Cambio'] != 0 else 0.0)
                    ),
                    axis=1
                ).round(1)
                
                evolucion['Tendencia_Diaria'] = (evolucion['Cambio'] / evolucion['Dias']).round(2)
                
                # Unir métricas de pendientes con producción
                evolucion = evolucion.merge(
                    prod_promedio,
                    left_on='OPERADOR_NORM',
                    right_on=col_operador,
                    how='left'
                ).drop(columns=[col_operador])
                
                # Calcular eficiencia (pendientes vs producción) con lógica mejorada y resaltado
                def calcular_eficiencia_v2(row):
                    eficiencia_real = row['Produccion_Promedio'] - row['Tendencia_Diaria'] # Cuánto reduce neto
                    produccion_promedio_minima_alta = 5
                    produccion_promedio_minima_media = 3
                    aumento_peligroso_pendientes = 1

                    if eficiencia_real > 0 and row['Produccion_Promedio'] >= produccion_promedio_minima_alta:
                        return 'Muy Alta'
                    elif eficiencia_real > 0 and row['Produccion_Promedio'] >= produccion_promedio_minima_media:
                        return 'Alta'
                    elif eficiencia_real > 0: # Producción baja pero reduce neto
                        return 'Mejorando'
                    elif eficiencia_real == 0 and row['Produccion_Promedio'] > 0: # Se mantiene, pero produce
                        return 'Estable'
                    # Casos donde los pendientes aumentan (eficiencia_real <= 0)
                    # Si la tendencia diaria es de aumento peligroso y la producción no la cubre
                    elif row['Tendencia_Diaria'] > aumento_peligroso_pendientes and row['Produccion_Promedio'] < row['Tendencia_Diaria']:
                        return 'En Observación' # ESTOS SE MARCARÁN EN ROJO
                    elif eficiencia_real < 0 and row['Produccion_Promedio'] > 0 and row['Produccion_Promedio'] > abs(row['Tendencia_Diaria']):
                        # Este caso es teóricamente cubierto por los primeros if, pero se deja por si acaso.
                        # Si produce más que el aumento absoluto de pendientes, pero el neto es negativo (implicaría que Tendencia_Diaria es negativa, contradicción)
                        # Lo reinterpreto: produce, pero no lo suficiente para que el neto sea positivo, pero sí más que el aumento absoluto de pendientes.
                        # Esto podría ser 'Conteniendo' si Tendencia_Diaria es positiva.
                        return 'Conteniendo'
                    elif eficiencia_real < 0 and row['Produccion_Promedio'] > 0:
                        return 'Conteniendo' # Produce, pero no lo suficiente para bajar pendientes netos.
                    else: # No produce o produce muy poco y los pendientes aumentan o no bajan.
                        return 'Baja'

                evolucion['Eficiencia'] = evolucion.apply(calcular_eficiencia_v2, axis=1)
                
                # Ordenar por tendencia diaria y luego por eficiencia para visualización
                # evolucion = evolucion.sort_values(by=['Tendencia_Diaria', 'Eficiencia'], ascending=[False, True])

                # Mostrar ranking con formato condicional
                def resaltar_criticos(row):
                    color = 'red' if row['Eficiencia'] == 'En Observación' else ''
                    return [f'background-color: {color}'] * len(row)

                st.dataframe(evolucion.style.apply(resaltar_criticos, axis=1), use_container_width=True)
                
                # Gráfico de dispersión
                fig_scatter = px.scatter(
                    evolucion,
                    x='Produccion_Promedio',
                    y='Tendencia_Diaria',
                    hover_name='OPERADOR_NORM',
                    color='Eficiencia',
                    size='Pendientes_Final',
                    title='Tendencia vs Producción Promedio',
                    labels={
                        'Produccion_Promedio': 'Producción Diaria Promedio',
                        'Tendencia_Diaria': 'Tendencia Diaria (pendientes/día)'
                    }
                )
                st.plotly_chart(fig_scatter, use_container_width=True)
                
                # Mostrar resumen por estado
                st.subheader("Resumen por Estado")
                resumen_estado = evolucion.groupby('Eficiencia').agg({
                    'OPERADOR_NORM': 'count',
                    'Pendientes_Final': 'sum',
                    'Produccion_Promedio': 'mean'
                }).round(2)
                st.dataframe(resumen_estado, use_container_width=True)
            else:
                st.warning("No hay datos suficientes para mostrar el ranking.")
        else:
            st.warning("No hay datos disponibles para el periodo seleccionado.")

except Exception as e:
    st.error(f"Error al cargar los datos: {str(e)}") 